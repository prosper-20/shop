import datetime
from django.utils import timezone
from django.core.management.base import BaseCommand
from customer.models import Customer
from openpyxl import load_workbook

class Command(BaseCommand):
    help = 'Populate Customer model from an Excel file'

    def handle(self, *args, **kwargs):
        # Path to the uploaded file (adjust as needed)
        file_path = "C:\\Users\\edwar\\Documents\\PROJECTS\\KLINSMAN_BROTHER\\MyProject\\new_NSI_Customer_Data_Requirements.xlsx"
        
        # Load the workbook
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active  # Assuming data is in the first sheet
        
        # Iterate over the rows in the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            try:
                Customer.objects.create(
    no=row[0] if len(row) > 0 else None,
    title=row[1] if len(row) > 1 else "Mr",
    name=row[2] if len(row) > 2 else "Unknown",
    business=row[3] if len(row) > 3 else "Unknown",
    email=row[4] if len(row) > 4 else "unknown@example.com",
    phone=row[5] if len(row) > 5 else "0000000000",
    dob=timezone.now().date(),
    address=row[7] if len(row) > 7 else "Unknown",
    state=row[8] if len(row) > 8 else "Others",
    other_state=row[9] if len(row) > 9 else None,
    occupation=row[10] if len(row) > 10 else "Unknown",
    nature=row[11] if len(row) > 11 else "Others",
    status=row[12] if len(row) > 12 else "new",
    is_reviewed=row[13] if len(row) > 13 else "Not Reviewed",
    date=timezone.now().date(),
    outstanding_balance=row[15] if len(row) > 15 else 0,
    data_entry_officer_note=row[16] if len(row) > 16 else None,
    review_officer_note=row[17] if len(row) > 17 else None,
    approval_officer_note=row[18] if len(row) > 18 else None,
    approval=bool(row[19]) if len(row) > 19 else False,
    exitdate=datetime.datetime.strptime(row[20], "%Y-%m-%d").date() if len(row) > 20 and row[20] else None,
    nextdue=datetime.datetime.strptime(row[21], "%Y-%m-%d").date() if len(row) > 21 and row[21] else None,
)

                self.stdout.write(f"Successfully added customer: {row[2]}")
            except Exception as e:
                self.stdout.write(f"Error adding customer {row[2]}: {str(e)}")
